#安裝套件:python excel (py+找尋套件名)
#在用第三方套件時，首先會先先建立一個物件，在存成簡寫(wb)，在使用物件簡寫(wd)的功能(用'.')
from openpyxl import Workbook
wb = Workbook() #type:Workbook，發明型別要寫class

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42
ws['B1'] = 'jane'
# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now() #datetime(套件).(的)datetime(模組).(的)now(功能)

# Save the file
wb.save("sample.xlsx")  #save存檔功能